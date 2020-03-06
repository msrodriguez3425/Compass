import requests
import pymongo
import pandas as pd
import time
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
from splinter import Browser
import win32com.client
from IPython.core.debugger import set_trace

def click_on_something(strings,types_,by_num = '',by_text = '',used_already = [],truncated = False):
    found = False
    my_dict = {}
    my_dict["strings"] = strings
    my_dict["types"] = types_
    #print(my_dict)
    j = 0
    while found == False:
        try: 
            to_click = browser
            for i in range(0,len(my_dict["strings"])):
                if my_dict["types"][i] == "id":
                    to_click = to_click.find_by_id(my_dict["strings"][i])
                elif my_dict["types"][i] == "css":
                    blah = f".{my_dict['strings'][i]}"
                    to_click = to_click.find_by_css(blah)
                    #In the case of selecting boundlists, must select the one that hasent been used yet
                    if used_already and (my_dict['strings'][i] == 'x-boundlist' or my_dict["strings"][i] == "list-ct"):
                        for obj in to_click:
                            if obj not in used_already:
                                to_click = obj
                else:
                    print("types_ invalid input: must be id or css")
                #print(f"to click: {to_click}")
            #If not looking to click on the first returned object
            if not by_num:
                if by_text:
                    #if searching element by text (e.g an option in a drop down)
                    if truncated == False:
                        clicked = False
                        for obj in to_click:
                            if obj.text == by_text:
                                obj.click()
                                clicked = True
                                break
                        if clicked == False:
                            print("by_text invalid input: text not found")
                            break
                    #if looking to click on element beginning with a specified word eg. MILLONES de pesos
                    elif truncated == True:
                        clicked = False
                        for obj in to_click:
                            if obj.text.split(" ")[0] == by_text:
                                obj.click()
                                clicked = True
                                break
                        if clicked == False:
                            print("by_text invalid input: text not found")
                            break
                    else:
                        print("truncated invalid input: must be True or False")
                #if not searching by text (e.g a drop down)
                else:
                    to_click.click()
                found = True
            #if looking to click on the first returned object
            elif by_num:
                to_click[int(by_num)].click()
                found = True
            else:
                print("by_num invalid input: must be empty or an integer")
                break
            return to_click
        
        except AttributeError:
            j += 1
            print(f"looking for element. attempt number {j}")
            
        if j > 10:
            print("Either app didn't load properly, or internet is down/weak")
            break

def scrape(num_,browser,app_dict):
    browser.visit(f'{base_url}{app_dict["url"][num_]}')    

    #Clicking the serie drop_down
    loaded = False
    while loaded == False:
        try:
            click_on_something(["tab-1013"],["id"])
            try:
                series_drop = click_on_something(["cmp_series_SERIE-triggerWrap","x-trigger-index-0"],["id","css"])
                loaded = True
            except:
                click_on_something(["tab-1012"],["id"])
                print("walao")
        except:
            pass

    #Clicking the todas las series button
    #select_all = browser.find_by_css(".list-ct").find_by_css(".x-column-header-text").first
    #select_all.click()
    series_clicked = False
    while series_clicked == False:
        print(f'test{len(browser.find_by_css(".list-ct"))}')
        click_on_something(["list-ct","x-column-header-text"],["css","css"],0)
        test = browser.find_by_id("cmp_series_SERIE-inputEl")._element.get_attribute('value')
        if test:
            series_clicked = True

    #clicking the serie drop_down
    unclicked = False
    while unclicked == False:
        try:
            series_drop.click()
            unclicked = True
        except:
            pass

    used_already = []
    used_already = add_to_used(used_already,browser.find_by_css(".list-ct"))
    used_already

    #This if is for tables in which there exists the option to select titulos
    if app_dict["titulos"][num_]:
        #Clicking the Títulos drop down
        tit_drop_clicked = False
        while tit_drop_clicked == False:
            try:
                titulos_drop = click_on_something(["cmp_serie_matriz-triggerWrap","x-trigger-index-0"],["id","css"])
                tit_drop_clicked = True
            except:
                pass

        used_already = add_to_used(used_already,browser.find_by_css(".list-ct"))

        #Clicking the todos los titulos button
        titulos_clicked = False
        while titulos_clicked == False:
            click_on_something(["list-ct","x-column-header-text"],["css","css"],0,'',used_already)
            test = browser.find_by_id("cmp_serie_matriz-inputEl")._element.get_attribute('value')
            if test:
                titulos_clicked = True

        #Unclicking the titulos drop_down
        unclicked = False
        while unclicked == False:
            try:
                titulos_drop.click()
                unclicked = True
            except:
                pass

    #clicking cifras drop_down
    cifras_drop_clicked = False
    while cifras_drop_clicked == False:
        try:
            cifras_drop = click_on_something(["cmp_cuadro_PRESENTACION","x-trigger-index-0"],["id","css"])
            cifras_drop_clicked = True
        except:
            pass    

    #Clicking on millones
    cifras_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],False,"Millones",used_already,True)

    #Upon clicking the dropdown, an x-boundlist is generated, must add it to used_aleady to we dont use it again
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year lower limit drop down
    #low_lim_drop = browser.find_by_id("cmp_cuadro_DE").find_by_css(".x-trigger-cell").find_by_css(".x-trigger-index-0")
    #low_lim_drop.click()
    low_lim_clicked = False
    while low_lim_clicked == False:
        try:
            low_lim_drop = click_on_something(["cmp_cuadro_DE","x-trigger-cell","x-trigger-index-0"],["id","css","css"])
            low_lim_clicked = True
        except:
            print("not found yet...")

    #Clicking on the second to latest year (hence the 1 in the click_on_something function) in the low_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],1,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year upper limit drop down
    #up_lim_drop = browser.find_by_id("cmp_cuadro_A").find_by_css(".x-trigger-index-0")
    #up_lim_drop.click()
    up_lim_clicked = False
    while up_lim_clicked == False:
        try:
            up_lim_drop = click_on_something(["cmp_cuadro_A","x-trigger-index-0"],["id","css"])
            up_lim_clicked = True
        except:
            print("not found yet...")

    #Clicking on the latest year in the up_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],0,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Consultar Series (clicking the button and suppressing warnings until the page loads the data)
    #consultar_series = browser.find_by_id("cmp_series_CONSULTAR_SERIES-btnEl")
    #consultar_series.click()
    success = False
    while success == False:
        clicked = False
        try:
            consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
            clicked = True
            print("consultar button clicked")
        except:
            print("no consultar button")
        if clicked == True:
            try:
                alert = browser.get_alert()
                alert.accept()
                print("alert accepted")
            except:
                print("there is no alert")
                time.sleep(0.5)
                test = browser.find_by_css(".x-mask-msg")
                #Sometimes the button isnt clicked, thus...
                print(f'test = {test}')
                if test == []:
                    consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
                    print("had to click again")
                success = True

    loaded = False
    while loaded == False:
        try:
            the_table = browser.find_by_id("GridSeries")
            header_row = the_table.find_by_css(".x-box-inner")[1]
            loaded = True
        except:
            print("loading...")
            #sometimes the consultar button isnt clicked in the previous step, thus...

    header_row.text

    #Finding Headers of the table
    header_divs = header_row.find_by_css(".x-unselectable")
    idx = 0
    headers = []
    for header in header_divs:
        if idx == 0:
            headers.append(header.find_by_tag("span").find_by_tag("label")._element.get_attribute("innerHTML").strip())
            idx += 1
        else:
            headers.append(header.find_by_tag("span")._element.get_attribute("innerHTML").replace("\n","").replace("\r","").strip())

    #Scraping the data and creating a data frame
    table_large = the_table.find_by_id("GridSeries-body")

    table = table_large.find_by_tag("table")

    rows = table.find_by_tag("tr")

    data_rows = list(rows)[1:]

    the_dict = dict((el,[]) for el in headers)

    for row in data_rows:
        to_append = row.text.split("\n")
        for k in range(0,len(the_dict)):
            the_dict[headers[k]].append(to_append[k]) 

    df =pd.DataFrame(the_dict)

    return df

def store(df_dict):
    #Loading the excel file
    print("Loading test2.xlsm...")
    wb = load_workbook(filename='N:/Lsolis/eco/SHCP/Historicos/test2.xlsm', read_only=False, keep_vba=True)
    idx = 0
    plsDownloadManually = []
    for worksheet in df_dict["sheet"]:
        #Loading the sheet
        time.sleep(1)
        sheet_ranges = wb[worksheet]
        print(worksheet)
        
        done = False
        row_ = 3
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                #print(sheet_ranges[f'A{row_}'].value)
                row_ += 1
            else:
                done = True
                
        last_row = row_ - 1
        #current_year is the year of the latest data already archived
        current_year = int(sheet_ranges[f'A{last_row}'].value[-4:])

        #current_month is the month of the latest data already archived
        #Here current_trim is defined incase this table is in quarters instead of months
        current_month = int(sheet_ranges[f'A{last_row}'].value[:2])
        current_trim = current_month
        
        #Code below takes care of case where data is being scraped during last month
        #  of the year (12) or last quarter of the year (04)

        #current_year_scraped is the year of the  latest data that was just scraped
        current_year_scraped = int(df_dict["df"][idx].iloc[-1,0].split("/")[1])

        if "Mensual" in df_dict["df"][idx].columns[0]:
            if current_month == 12 and current_year_scraped > current_year:
                current_year += 1
        elif "Trimestral" in df_dict["df"][idx].columns[0]:
            if current_trim == 4 and current_year_scraped > current_year:
                current_year += 1
        else:
            print("error with time headers")
            break

        if current_year == current_year_scraped:
            previous_2year = current_year - 2
        elif (current_year_scraped - current_year) == 1:
            previous_2year = current_year  - 1
        else:
            print("Please download this table by hand from SHCP, replace in test2, and try again.")
            break

        #This finds the number of rows of data so far in the 2 years
        done = False
        row_ = 4
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                if int(sheet_ranges[f'A{row_}'].value[-4:]) == previous_2year:
                    last_row_2previous = row_
            #print(sheet_ranges[f'A{row_}'].value)  
                row_ += 1
            else:
                done = True


                
        #insert rows if there are rows to be inserted        
        if len(df_dict["df"][idx]) > (last_row-last_row_2previous):
            sheet_ranges.insert_rows(last_row + 1, len(df_dict["df"][idx])-(last_row - last_row_2previous))
            print(f'{len(df_dict["df"][idx]) - (last_row-last_row_2previous)} rows were inserted.')

        #finding row range to which values will be inserted
        ins_begin = last_row_2previous + 1
        ins_end = ins_begin + len(df_dict["df"][idx]) - 1
        


        #Making sure headers in excel are equal to headers in dataframe
        sheet_col_heads = []
        sheet_col_heads2 = []
        for col in range(0,count_columns(sheet_ranges)):
            #print(count_columns(sheet_ranges))
            sheet_col_heads.append(sheet_ranges['A3'].offset(0,col).value.replace("_x000D_",""))
            sheet_col_heads2.append(sheet_ranges['A3'].offset(0,col).value)

        stripped_col_heads = []
        for head in sheet_col_heads:
            stripped_col_heads.append(head.strip())

        stripped_col_heads2 = []
        for head in sheet_col_heads2:
            stripped_col_heads2.append(head.strip())

        compare = list(df_dict["df"][idx].columns)
        compare2 = []

        for header in compare:
            compare2.append(header.replace("\r","").replace("\n","").strip())
      
        
        print(f'{df_dict["sheet"][idx]}: {stripped_col_heads == compare2}')
        #comparing the sheet columns to the scraped columns
        if stripped_col_heads == compare2:
            headers_agree = True
        else:
            plsDownloadManually.append(worksheet)
            headers_agree = False
            print(f"excel sheet column headers don't agree with scraped column headers. Download {worksheet} by hand and replace in test2.xlsm with exact same format as other tables (column numbers on top, headers on row 3).")

        if headers_agree:
            #Insert Values from dataframe into excel with some format
            row_height = sheet_ranges.row_dimensions[last_row_2previous].height
            for col in range(1,len(stripped_col_heads) + 1):
                for rw in range(ins_begin, ins_end + 1):
                    #pasting value in cell
                    try:
                        #this try for numeric data
                        sheet_ranges.cell(rw,col).value = float((df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",",""))
                        sheet_ranges.cell(rw,col).number_format = '#,###.0'
                    except ValueError:
                        #this try for dates
                        sheet_ranges.cell(rw,col).value = (df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",","")
                    #setting font style and size    
                    sheet_ranges.cell(rw,col).font = Font(name='Arial', size=9)
                    #setting border of cell
                    border = Border(left=Side(border_style='thin', color='E4E4E4'),
                                    right=Side(border_style='thin', color='E4E4E4'),
                                    top=Side(border_style='thin', color='E4E4E4'),
                                    bottom=Side(border_style='thin', color='E4E4E4'))
                    sheet_ranges.cell(rw,col).border = border
                    #setting alignment of text in cell
                    alignment=Alignment(horizontal='right',
                                        vertical='top')
                    sheet_ranges.cell(rw,col).alignment = alignment
                    #setting row height
                    sheet_ranges.row_dimensions[rw].height = row_height
                    #setting fill color of cell
                    greyFill = PatternFill(start_color='DCDCDC',
                                end_color='DCDCDC',
                                fill_type='solid')
                    whiteFill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')
                    #Grey/White/Grey/White
                    if rw % 2 == 0:
                        sheet_ranges.cell(rw,col).fill = greyFill 
                    else:
                        sheet_ranges.cell(rw,col).fill = whiteFill
              
                print(f"column {col} out of {len(stripped_col_heads)} inserted successfully")

        idx += 1
    print("Saving test2.xlsm...")
    wb.save(filename = 'N:/Lsolis/eco/SHCP/Historicos/test2.xlsm')
    print("Data was stored in test2.")
    return plsDownloadManually

def add_to_used(used_already,xboundlists):
    for lst in xboundlists:
        el_id = lst._element.get_attribute('id')
        if el_id not in used_already:
            used_already.append(el_id)
    return used_already

def count_columns(sheet_ranges):
    non_empty = True
    col_ = 1
    col_count = 0
    while non_empty:
        if sheet_ranges['A3'].offset(0,col_ - 1).value:
            col_ +=1
            col_count += 1
        else:
            non_empty = False
    return col_count

def scraped_to_official():
    
    print("Copying data from test2 to official file...")
    xl=win32com.client.Dispatch("Excel.Application")
    
    print(" Loading official file...")
    wb = xl.Workbooks.Open("N:/Lsolis/eco/SHCP/Historicos/2Deuda_publica.xlsm")
    
    print(" Running macro...")
    xl.Application.Run("2Deuda_publica.xlsm!CopySheetsFromScrapedData")

    wb.Save()
    wb.Close()
    del xl
    print("Finished!")
              
    

app_dict = {
    "title":[
        "Deuda Interna del Sector Público Federal",
        "Deuda Externa del Sector Público Federal",
        "Saldos de la Deuda del Sector Público Federal",
        "Saldos de la Deuda Pública Externa por Deudor Directo ante el Extranjero y Usuario de Recursos",
        "Saldos de la Deuda Pública Externa Clasificada por País y Moneda",
        "Evolución del Endeudamiento Externo del Sector Público Federal",
        "Colocaciones en los Mercados Internacionales",
        "Deuda Interna del Gobierno Federal",
        "Deuda Externa del Gobierno Federal",
        "Saldos de la Deuda del Gobierno Federal",
        "Saldo de las Obligaciones Garantizadas del Gobierno Federal"
    ],
    "url":[
        "Deuda%20Interna%20del%20Sector%20P%C3%BAblico%20Federal&param_formato=1&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Deuda%20Externa%20del%20Sector%20P%C3%BAblico%20Federal&param_formato=2&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Saldos%20de%20la%20Deuda%20del%20Sector%20P%C3%BAblico%20Federal&param_formato=3&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Saldos%20de%20la%20Deuda%20P%C3%BAblica%20Externa%20por%20Deudor%20Directo%20ante%20el%20Extranjero%20y%20Usuario%20de%20Recursos&param_formato=6&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Saldos%20de%20la%20Deuda%20P%C3%BAblica%20Externa%20Clasificada%20por%20Pa%C3%ADs%20y%20Moneda&param_formato=7&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Evoluci%C3%B3n%20del%20Endeudamiento%20Externo%20del%20Sector%20P%C3%BAblico%20Federal&param_formato=8&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Colocaciones%20en%20los%20Mercados%20Internacionales&param_formato=5&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=7&",
        "Deuda%20Interna%20del%20Gobierno%20Federal&param_formato=1&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=6&",
        "Deuda%20Externa%20del%20Gobierno%20Federal&param_formato=2&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=6&",
        "Saldos%20de%20la%20Deuda%20del%20Gobierno%20Federal&param_formato=3&param_unidad=1&param_tipo=2&param_lenguaje=1&param_clasificacion=6&",
        "Saldo%20de%20las%20Obligaciones%20Garantizadas%20del%20Gobierno%20Federal&param_formato=1&param_unidad=2&param_tipo=5&param_lenguaje=1&param_clasificacion=21&"
    ],
    "sheet":[
        "2.1.1S",
        "2.1.2S",
        "2.1.3S",
        "2.1.4S",
        "2.1.5S",
        "2.1.6S",
        "2.1.7S",
        "2.2.1S",
        "2.2.2S",
        "2.2.3S",
        "2.3.1S"
    ],
    "titulos":[
        True,
        True,
        False,
        False,
        False,
        False,
        False,
        True,
        True,
        False,
        False
    ]
}

executable_path = {'executable_path': 'chromedriver.exe'}
browser = Browser('chrome', **executable_path, headless=False)
base_url = "http://presto.hacienda.gob.mx/presto/files/system/mashlets/app_layout_estopor/index.html?param_formato_desc="
#table_url = "/presto/files/system/mashlets/app_layout_estopor/index.html?param_formato_desc=Balance%20del%20Sector%20P%C3%BAblico&param_formato=1&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1& "

#url = "/presto/files/system/mashlets/app_layout_estopor/index.html?param_formato_desc=Balance%20del%20Sector%20P%C3%BAblico&param_formato=1&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&"
#url = "http://www.banxico.org.mx/"

df_dict = {
    "title":[],
    "df":[],
    "sheet":[]
}

for i in range(0,len(app_dict["title"])):
    df = scrape(i,browser,app_dict)
    title = browser.find_by_css(".cls-title-table").find_by_css(".cls-title-main").text
    df_dict["title"].append(title)
    df_dict["df"].append(df)
    df_dict["sheet"].append(app_dict["sheet"][i])
    
to_download = store(df_dict)
              
scraped_to_official()

if to_download:
    print("\n\n\n")
    print("IMPORTANT: please download the following tables by hand...")
    print("\n")
    for table in to_download:
        print(f'--- {table}')
    print("\n")
    print(f"SHCP inserted a new column in these tables. Download by hand and replace in test2.xlsm with exact same format as other tables (column numbers on top, headers on row 3), save and close test2.xlsm, and execute this script again. At this time, the rest of the tables have been updated as normal.")
    print("\n")
    print("To download a table(s): \n ")
    print("-- Go to http://www.shcp.gob.mx/POLITICAFINANCIERA/FINANZASPUBLICAS/Estadisticas_Oportunas_Finanzas_Publicas/Paginas/unica2.aspx")
    print("-- Open needed table: example 2.1.3S -- > 1 means first tab in 2nd row i.e. Deuda Pública, 1 means 1st tab of 3rd row i.e. Deuda del Sector Público, 3 means 3rd available table i.e. Saldos de la Deuda del Sector Público Federal.")
    print("-- Select Series tab")
    print("-- Select Series drop down, select Todas las series")
    print("-- If there's a Títulos drop down, click on it and select Todos los titulos")
    print("-- Select Cifras drop down, select Millones de pesos")
    print("-- Select De drop down, select earliest year available")
    print("-- Select A drop down, select latest year available")
    print("-- Click Consultar Series")
    print("-- Click Exportar Excel")
    print("-- Open downloaded file")
    print("-- Insert row on top, number columns (starting with 1 at column A until last column)")
    print("-- Replace in test2.xlsm with the appropriate sheet name i.e. 2.1.3S")
    print("-- Save and close test2.xlsm")  