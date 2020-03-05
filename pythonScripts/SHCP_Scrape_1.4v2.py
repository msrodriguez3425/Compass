"""
Created on Mon Feb 10 17:40:08 2020

@author: mrodriguezo
"""

import requests
import pymongo
import pandas as pd
import time
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
from splinter import Browser
import win32com.client

def click_on_something(strings,types_,by_num = '',by_text = '',used_already = [],truncated = False):
    found = False
    my_dict = {}
    my_dict["strings"] = strings
    my_dict["types"] = types_
    #print(my_dict)
    j = 0
    while found == False:
        try: 
            to_click = browser
            for i in range(0,len(my_dict["strings"])):
                if my_dict["types"][i] == "id":
                    to_click = to_click.find_by_id(my_dict["strings"][i])
                elif my_dict["types"][i] == "css":
                    blah = f".{my_dict['strings'][i]}"
                    to_click = to_click.find_by_css(blah)
                    #In the case of selecting boundlists, must select the one that hasent been used yet
                    if used_already and (my_dict['strings'][i] == 'x-boundlist' or my_dict["strings"][i] == "list-ct"):
                        for obj in to_click:
                            if obj not in used_already:
                                to_click = obj
                else:
                    print("types_ invalid input: must be id or css")
                    break
                #print(f"to click: {to_click}")
            #If not looking to click on the first returned object
            if not by_num:
                if by_text:
                    clicked = False
                    for obj in to_click:
                        if obj.text == by_text:
                            obj.click()
                            clicked = True
                            break
                    if clicked == False:
                        print("by_text invalid input: text not found")
                        break
                else:
                    to_click.click()
                found = True
            #if looking to click an object by number
            elif by_num:
                to_click[int(by_num)].click()
                found = True
            else:
                print("by_num invalid input: must be empty or an integer")
                break
            return to_click
        
        except AttributeError:
            j += 1
            print(f"looking for element. attempt number {j}")
            
        if j > 10:
            print("Either app didn't load properly, or internet is down/weak")
            break

def scrape(num_,browser,app_dict):
    browser.visit(f'{base_url}{app_dict["url"][num_]}')
    
    #Clicking the serie drop_down
    loaded = False
    while loaded == False:
        try:
            click_on_something(["tab-1013"],["id"])
            try:
                series_drop = click_on_something(["cmp_series_SERIE-triggerWrap","x-trigger-index-0"],["id","css"])
                loaded = True
            except:
                click_on_something(["tab-1012"],["id"])
                print("walao")
        except:
            pass
        
    #Clicking the todas las series button
    #select_all = browser.find_by_css(".list-ct").find_by_css(".x-column-header-text").first
    #select_all.click()
    series_clicked = False
    while series_clicked == False:
        #print(f'test{len(browser.find_by_css(".list-ct"))}')
        click_on_something(["list-ct","x-column-header-text"],["css","css"],0)
        test = browser.find_by_id("cmp_series_SERIE-inputEl")._element.get_attribute('value')
        if test:
            series_clicked = True
        

    #unclicking the serie drop_down
    unclicked = False
    while unclicked == False:
        try:
            series_drop.click()
            unclicked = True
        except:
            pass

    used_already = []
    used_already = add_to_used(used_already,browser.find_by_css(".list-ct"))
    used_already
            
        
    if app_dict["titulos"][i]:
        #Clicking the Títulos drop down
        tit_drop_clicked = False
        while tit_drop_clicked == False:
            try:
                titulos_drop = click_on_something(["cmp_serie_matriz-triggerWrap","x-trigger-index-0"],["id","css"])
                tit_drop_clicked = True
            except:
                pass

        used_already = add_to_used(used_already,browser.find_by_css(".list-ct"))
        used_already

        #Clicking the todos los titulos button
        titulos_clicked = False
        while titulos_clicked == False:
            click_on_something(["list-ct","x-column-header-text"],["css","css"],True,'',used_already)
            test = browser.find_by_id("cmp_serie_matriz-inputEl")._element.get_attribute('value')
            if test:
                titulos_clicked = True

        #unclicking the titulos drop_down
        unclicked = False
        while unclicked == False:
            try:
                titulos_drop.click()
                unclicked = True
            except:
                pass
        
    #clicking cifras drop_down
    cifras_drop_clicked = False
    while cifras_drop_clicked == False:
        try:
            cifras_drop = click_on_something(["cmp_cuadro_PRESENTACION","x-trigger-index-0"],["id","css"])
            cifras_drop_clicked = True
        except:
            print("cifras drop not found yet...")

    #Clicking on millones
    cifras_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],False,"Millones",used_already,True)



    #Upon clicking the dropdown, an x-boundlist is generated, must add it to used_aleady to we dont use it again
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year lower limit drop down
    #low_lim_drop = browser.find_by_id("cmp_cuadro_DE").find_by_css(".x-trigger-cell").find_by_css(".x-trigger-index-0")
    #low_lim_drop.click()
    low_lim_clicked = False
    while low_lim_clicked == False:
        try:
            low_lim_drop = click_on_something(["cmp_cuadro_DE","x-trigger-cell","x-trigger-index-0"],["id","css","css"])
            low_lim_clicked = True
        except:
            print("year lower limit drop down not found yet...")

    #Clicking on the  second to latest year ( hence the 1 in the click_on_something function) in the low_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],1,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year upper limit drop down
    #up_lim_drop = browser.find_by_id("cmp_cuadro_A").find_by_css(".x-trigger-index-0")
    #up_lim_drop.click()
    up_lim_clicked = False
    while up_lim_clicked == False:
        try:
            up_lim_drop = click_on_something(["cmp_cuadro_A","x-trigger-index-0"],["id","css"])
            up_lim_clicked = True
        except:
            print("up_lim_drop not found yet...")

    #Clicking on the latest year in the up_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],0,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Consultar Series (clicking the button and suppressing warnings until the page loads the data)
    #consultar_series = browser.find_by_id("cmp_series_CONSULTAR_SERIES-btnEl")
    #consultar_series.click()
    success = False
    while success == False:
        clicked = False
        try:
            consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
            clicked = True
            print("consultar button clicked")
        except:
            print("no consultar button")
        if clicked == True:
            try:
                alert = browser.get_alert()
                alert.accept()
                print("alert accepted")
            except:
                print("there is no alert")
                time.sleep(0.5)
                test = browser.find_by_css(".x-mask-msg")
                #Sometimes the button isnt clicked, thus...
                print(f'test = {test}')
                if test == []:
                    consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
                    print("had to click again")
                success = True


    loaded = False
    while loaded == False:
        try:
            the_table = browser.find_by_id("GridSeries")
            header_row = the_table.find_by_css(".x-box-inner")[1]
            loaded = True
        except:
            print("loading...")
            #sometimes the consultar button isnt clicked in the previous step, thus...

    
    #Finding Headers of the table
    header_divs = header_row.find_by_css(".x-unselectable")
    idx = 0
    headers = []
    for header in header_divs:
        if idx == 0:
            headers.append(header.find_by_tag("span").find_by_tag("label")._element.get_attribute("innerHTML").strip())
            idx += 1
        else:
            headers.append(header.find_by_tag("span")._element.get_attribute("innerHTML").replace("\n","").replace("\r","").strip())

    #Scraping the data and creating a data frame
    table_large = the_table.find_by_id("GridSeries-body")

    table = table_large.find_by_tag("table")

    rows = table.find_by_tag("tr")

    data_rows = list(rows)[1:]

    the_dict = dict((el,[]) for el in headers)

    for row in data_rows:
        to_append = row.text.split("\n")
        for k in range(0,len(the_dict)):
            the_dict[headers[k]].append(to_append[k]) 
#
    df =pd.DataFrame(the_dict)
    return df

def store(df_dict):
    #Loading the excel file
    print("Loading test file...")
    wb = load_workbook(filename='N:/Lsolis/eco/SHCP/Historicos/test.xlsm', read_only=False, keep_vba=True)
    idx = 0
    plsDownloadManually = []
    for worksheet in df_dict["sheet"]:
        #Loading the sheet
        sheet_ranges = wb[worksheet]

        #determining whether or not a new row(s) must be inserted
        done = False
        row_ = 3
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                #print(sheet_ranges[f'A{row_}'].value)
                row_ += 1
            else:
                done = True
        last_row = row_ - 1
        current_year = int(sheet_ranges[f'A{last_row}'].value[-4:])

        current_month = int(sheet_ranges[f'A{last_row}'].value[:2])
        current_trim = current_month

        #this takes care of case where data is being scraped during last month of the year (12) or last quarter of the year (04)
        current_year_scraped = int(df_dict["df"][idx].iloc[-1,0].split("/")[1])
        if "Mensual" in df_dict["df"][idx].columns[0]:
            if current_month == 12 and current_year_scraped > current_year:
                current_year += 1
        elif "Trimestral" in df_dict["df"][idx].columns[0]:
            if current_trim == 4 and current_year_scraped > current_year:
                current_year += 1
        else:
            print("error with time headers")
            break

        if current_year == current_year_scraped:
            previous_2year = current_year - 2
        elif (current_year_scraped - current_year) == 1:
            previous_2year = current_year  - 1
        else:
            print("Please download this table by hand from SHCP, replace in test2, and try again.")
            break
    

        #This finds the number of rows of data so far in the 2 years
        done = False
        row_ = 4
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                if int(sheet_ranges[f'A{row_}'].value[-4:]) == previous_2year:
                    last_row_2previous = row_
            #print(sheet_ranges[f'A{row_}'].value)  
                row_ += 1
            else:
                done = True

        #insert rows if there are rows to be inserted        
        if len(df_dict["df"][idx]) > (last_row-last_row_2previous):
            sheet_ranges.insert_rows(last_row + 1, len(df_dict["df"][idx])-(last_row - last_2row_previous))
            print(f'{len(df_dict["df"][idx]) - (last_row-last_row_2previous)} rows were inserted.')

        #finding row range to which values will be inserted
        ins_begin = last_row_2previous + 1
        ins_end = ins_begin + len(df_dict["df"][idx]) - 1

        #Making sure headers in excel are equal to headers in dataframe
        sheet_col_heads = []
        sheet_col_heads2 = []
        for col in range(0,count_columns(sheet_ranges)):
            #print(count_columns(sheet_ranges))
            sheet_col_heads.append(sheet_ranges['A3'].offset(0,col).value.replace("_x000D_",""))
            sheet_col_heads2.append(sheet_ranges['A3'].offset(0,col).value)
            
        stripped_col_heads = []
        for head in sheet_col_heads:
            stripped_col_heads.append(head.strip())
            
        stripped_col_heads2 = []
        for head in sheet_col_heads2:
            stripped_col_heads2.append(head.strip())
            
        compare = list(df_dict["df"][idx].columns)
        compare2 = []

        for header in compare:
            compare2.append(header.replace("\r","").replace("\n","").strip())
            
        print(f"{worksheet}: {stripped_col_heads == compare2}")

        #comparing the sheet columns to the scraped columns
        if stripped_col_heads == compare2:
            headers_agree = True
        else:
            plsDownloadManually.append(worksheet)
            headers_agree = False
            print(f"excel sheet column headers don't agree with scraped column headers. Download {worksheet} by hand and replace in test.xlsm with exact same format as other tables (column numbers on top, headers on row 3).")
            
        if headers_agree:
            #Insert Values from dataframe into excel with some format
            row_height = sheet_ranges.row_dimensions[last_row_2previous].height
            for col in range(1,len(stripped_col_heads) + 1):
                for rw in range(ins_begin, ins_end + 1):
                    #pasting value in cell
                    try:
                        sheet_ranges.cell(rw,col).value = float((df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",",""))
                        sheet_ranges.cell(rw,col).number_format = '#,###.0'
                    except ValueError:
                        sheet_ranges.cell(rw,col).value = (df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",","")
                    #setting font style and size    
                    sheet_ranges.cell(rw,col).font = Font(name='Arial', size=9)
                    #setting border of cell
                    border = Border(left=Side(border_style='thin', color='E4E4E4'),
                                    right=Side(border_style='thin', color='E4E4E4'),
                                    top=Side(border_style='thin', color='E4E4E4'),
                                    bottom=Side(border_style='thin', color='E4E4E4'))
                    sheet_ranges.cell(rw,col).border = border
                    #setting alignment of text in cell
                    alignment=Alignment(horizontal='right',
                                        vertical='top')
                    sheet_ranges.cell(rw,col).alignment = alignment
                    #setting row height
                    sheet_ranges.row_dimensions[rw].height = row_height
                    #setting fill color of cell
                    greyFill = PatternFill(start_color='DCDCDC',
                                end_color='DCDCDC',
                                fill_type='solid')
                    whiteFill = PatternFill(start_color='FFFFFF',
                        end_color='FFFFFF',
                        fill_type='solid')
                    #Grey/White/Grey/White
                    if rw % 2 == 0:
                        sheet_ranges.cell(rw,col).fill = greyFill 
                    else:
                        sheet_ranges.cell(rw,col).fill = whiteFill
                print(f"column {col} out of {len(stripped_col_heads)} inserted successfully")
        idx += 1

    wb.save(filename = 'N:/Lsolis/eco/SHCP/Historicos/test.xlsm')
    print("test.xlsm saved")
    return plsDownloadManually

def add_to_used(used_already,xboundlists):
    for lst in xboundlists:
        el_id = lst._element.get_attribute('id')
        if el_id not in used_already:
            used_already.append(el_id)
    return used_already

def count_columns(sheet_ranges):
    non_empty = True
    col_ = 1
    col_count = 0
    while non_empty:
        if sheet_ranges['A3'].offset(0,col_ - 1).value:
            col_ +=1
            col_count += 1
        else:
            non_empty = False
    return col_count

def scraped_to_official():
    print("Copying data from test to official file.")
    xl=win32com.client.Dispatch("Excel.Application")
    print(" Loading official file...")
    wb = xl.Workbooks.Open("N:/Lsolis/eco/SHCP/Historicos/1Ingreso_gasto_fincanc_sector_publico.xlsm")
    print(" Running macro...")
    xl.Application.Run("1Ingreso_gasto_fincanc_sector_publico.xlsm!CopySheets4FromScrapedData")

    wb.Save()
    wb.Close()
    del xl
    print("Finished!")

app_dict = {
    "title":[
        "I. Situación Financiera del Sector Público no Financiero",
        "II. Situación Financiera del Gobierno Federal",
        "III. Situación Financiera del Gobierno Federal y Seguridad Social",
        "Situación Financiera de Organismos y Empresas Bajo Control Presupuestario Directo. Consolidado",
        "Situación Financiera de Organismos y Empresas Bajo Control Presupuestario Directo. Consolidado sin PEMEX",
        "Situación Financiera de Pemex",
        "Situación Financiera de CFE",
        "Situación Financiera de LFC",
        "Situación Financiera del IMSS",
        "Situación Financiera del ISSSTE",
        #"Situación Financiera de Otros Organismos y Empresas Bajo Control Presupuestario Directo (Histórico)",
        #Lo dejaron de actualizar en  2016
        "Situación Financiera de Organismos y Empresas Bajo Control Presupuestario Indirecto. Consolidado",
        "Situación Financiera de Pronósticos para la Asistencia Pública",
        "Situación Financiera de Lotería Nacional para la Asistencia Pública",
        "Situación Financiera del Instituto de Seguridad Social para las Fuerzas Armadas Mexicanas",
        "Situación Financiera del Fideicomiso de Riesgo Compartido",
        "Situación Financiera de Aeropuertos y Servicios Auxiliares",
        "Situación Financiera de Telecomunicaciones de México",
        "Situación Financiera del Servicio Postal Mexicano",
        "Situación Financiera de Ferrocarriles Nacionales de México",
        "Situación Financiera de Caminos y Puentes Federales de Ingresos y Servicios Conexos",
        "Situación Financiera del Instituto Nacional para la Educación de los Adultos",
        "Situación Financiera del Consejo Nacional de Ciencia y Tecnología",
        "Situación Financiera del Colegio Nacional de Educación Profesional Técnica",
        "Situación Financiera de la Comisión Nacional de Libros de Texto Gratuitos",
        "Situación Financiera del Comité Administrador del Programa Federal de Construcción de Escuelas",
        "Situación Financiera del Consejo Nacional de Fomento Educativo",
        "Situación Financiera del Hospital General de México",
        "Situación Financiera del Sistema Nacional Para el Desarrollo Integral de la Familia",
        "Situación Financiera del Instituto Mexicano del Petróleo",
        "Situación Financiera de P.M.I. Comercio Internacional",
        "Situación Financiera de DICONSA",
        "Situación Financiera de LICONSA",
        #"Situación Financiera de BORUCONSA",
        #Lo dejaron de actualizar en 2016
        #"Situación Financiera del Metro",
        #Lo dejaron de actualizar en 2016
        #"Situación Financiera de Telmex",
        #Lo dejaron de actualizar en 2016
        #"Situación Financiera de AHMSA",
        #Lo dejaron de actualizar en 2016
        #"Situación Financiera del DDF",
        #Lo dejaron de actualizar en 2016
        "Situación Financiera de Otros Organismos y Empresas Bajo Control Presupuestario Indirecto",
        "Consolidado Total de Bancos de Desarrollo y Fondos y Fideicomisos",
        "Consolidado de Bancos de Desarrollo",
        "Nacional Financiera, S.N.C. (NAFIN)",
        "Banco Nacional de Obras y Servicios, S.N.C. (BANOBRAS)",
        "Banco Nacional de Comercio Exterior, S.N.C. (BANCOMEXT)",
        "Banco del Ahorro Nacional y Servicios Financieros (BANSEFI)",
        "Sociedad Hipotecaria Federal, S.N.C. (HIPOTECARIA)",
        "Banco Nacional del Ejército, Fuerza Aérea y Armada, S.N.C. (BANJERCITO)",
        "Otros Bancos (Histórico)",
        "Consolidado de Fondos y Fideicomisos",
        "FIRA (Fondo, Fefa y Fega)",
        "Fondo de Operación y Financiamiento Bancario a la Vivienda (FOVI)",
        "Fondo de Garantía y Fomento para las Actividades Pesqueras (FOPESCA)",
        "Fondo Nacional de Habitaciones Populares (FONHAPO)",
        "Instituto del Fondo Nacional para el Consumo de los Trabajadores (INFONACOT)",
        "Fideicomiso de Fomento Minero (FIFOMI)",
        "Fondo Nacional de Fomento al Turismo (FONATUR)",
        "Otros Fondos y Fideicomisos (Histórico)",
        "Financiera Nacional de Desarrollo Agropecuario, Rural, Forestal y Pesquero"
    ],
    "url":[
        "I.%20Situaci%C3%B3n%20Financiera%20del%20Sector%20P%C3%BAblico%20no%20Financiero&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=28&",
        "II.%20Situaci%C3%B3n%20Financiera%20del%20Gobierno%20Federal&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=2&",
        "III.%20Situaci%C3%B3n%20Financiera%20del%20Gobierno%20Federal%20y%20Seguridad%20Social&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=4&",
        "Situaci%C3%B3n%20Financiera%20de%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Directo.%20Consolidado&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20de%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Directo.%20Consolidado%20sin%20PEMEX&param_formato=2&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20de%20Pemex&param_formato=3&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20de%20CFE&param_formato=4&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20de%20LFC&param_formato=5&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20del%20IMSS&param_formato=6&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        "Situaci%C3%B3n%20Financiera%20del%20ISSSTE&param_formato=7&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        #"Situaci%C3%B3n%20Financiera%20de%20Otros%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Directo%20(Hist%C3%B3rico)&param_formato=8&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=30&",
        #Lo dejaron de actualizar en 2016
        "Situaci%C3%B3n%20Financiera%20de%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Indirecto.%20Consolidado&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Pron%C3%B3sticos%20para%20la%20Asistencia%20P%C3%BAblica&param_formato=2&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Loter%C3%ADa%20Nacional%20para%20la%20Asistencia%20P%C3%BAblica&param_formato=3&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Instituto%20de%20Seguridad%20Social%20para%20las%20Fuerzas%20Armadas%20Mexicanas&param_formato=4&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Fideicomiso%20de%20Riesgo%20Compartido&param_formato=5&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Aeropuertos%20y%20Servicios%20Auxiliares&param_formato=6&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Telecomunicaciones%20de%20M%C3%A9xico&param_formato=7&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Servicio%20Postal%20Mexicano&param_formato=8&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Ferrocarriles%20Nacionales%20de%20M%C3%A9xico&param_formato=9&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20Caminos%20y%20Puentes%20Federales%20de%20Ingresos%20y%20Servicios%20Conexos&param_formato=10&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Instituto%20Nacional%20para%20la%20Educaci%C3%B3n%20de%20los%20Adultos&param_formato=11&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Consejo%20Nacional%20de%20Ciencia%20y%20Tecnolog%C3%ADa&param_formato=12&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Colegio%20Nacional%20de%20Educaci%C3%B3n%20Profesional%20T%C3%A9cnica&param_formato=13&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20la%20Comisi%C3%B3n%20Nacional%20de%20Libros%20de%20Texto%20Gratuitos&param_formato=14&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Comit%C3%A9%20Administrador%20del%20Programa%20Federal%20de%20Construcci%C3%B3n%20de%20Escuelas&param_formato=15&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Consejo%20Nacional%20de%20Fomento%20Educativo&param_formato=16&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Hospital%20General%20de%20M%C3%A9xico&param_formato=17&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Sistema%20Nacional%20Para%20el%20Desarrollo%20Integral%20de%20la%20Familia&param_formato=18&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20del%20Instituto%20Mexicano%20del%20Petr%C3%B3leo&param_formato=19&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20P.M.I.%20Comercio%20Internacional&param_formato=20&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20DICONSA&param_formato=21&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Situaci%C3%B3n%20Financiera%20de%20LICONSA&param_formato=22&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #"Situaci%C3%B3n%20Financiera%20de%20BORUCONSA&param_formato=23&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #Lo dejaron de actualizar en 2016
        #"Situaci%C3%B3n%20Financiera%20del%20Metro&param_formato=24&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #Lo dejaron de actualizar en 2016
        #"Situaci%C3%B3n%20Financiera%20de%20Telmex&param_formato=25&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #Lo dejaron de actualizar en 2016
        #"Situaci%C3%B3n%20Financiera%20de%20AHMSA&param_formato=26&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #Lo dejaron de actualizar en 2016
        #"Situaci%C3%B3n%20Financiera%20del%20DDF&param_formato=27&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        #Lo dejaron de actualizar en 2016
        "Situaci%C3%B3n%20Financiera%20de%20Otros%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Indirecto&param_formato=28&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=31&",
        "Consolidado%20Total%20de%20Bancos%20de%20Desarrollo%20y%20Fondos%20y%20Fideicomisos&param_formato=1&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Consolidado%20de%20Bancos%20de%20Desarrollo&param_formato=2&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Nacional%20Financiera%2C%20S.N.C.%20(NAFIN)&param_formato=3&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Banco%20Nacional%20de%20Obras%20y%20Servicios%2C%20S.N.C.%20(BANOBRAS)&param_formato=4&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Banco%20Nacional%20de%20Comercio%20Exterior%2C%20S.N.C.%20(BANCOMEXT)&param_formato=5&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Banco%20del%20Ahorro%20Nacional%20y%20Servicios%20Financieros%20(BANSEFI)&param_formato=6&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Sociedad%20Hipotecaria%20Federal%2C%20S.N.C.%20(HIPOTECARIA)&param_formato=7&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Banco%20Nacional%20del%20Ej%C3%A9rcito%2C%20Fuerza%20A%C3%A9rea%20y%20Armada%2C%20S.N.C.%20(BANJERCITO)&param_formato=8&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Otros%20Bancos%20(Hist%C3%B3rico)&param_formato=9&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Consolidado%20de%20Fondos%20y%20Fideicomisos&param_formato=10&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "FIRA%20(Fondo%2C%20Fefa%20y%20Fega)&param_formato=11&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Fondo%20de%20Operaci%C3%B3n%20y%20Financiamiento%20Bancario%20a%20la%20Vivienda%20(FOVI)&param_formato=12&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Fondo%20de%20Garant%C3%ADa%20y%20Fomento%20para%20las%20Actividades%20Pesqueras%20(FOPESCA)&param_formato=13&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Fondo%20Nacional%20de%20Habitaciones%20Populares%20(FONHAPO)&param_formato=14&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Instituto%20del%20Fondo%20Nacional%20para%20el%20Consumo%20de%20los%20Trabajadores%20(INFONACOT)&param_formato=15&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Fideicomiso%20de%20Fomento%20Minero%20(FIFOMI)&param_formato=16&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Fondo%20Nacional%20de%20Fomento%20al%20Turismo%20(FONATUR)&param_formato=17&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Otros%20Fondos%20y%20Fideicomisos%20(Hist%C3%B3rico)&param_formato=18&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&",
        "Financiera%20Nacional%20de%20Desarrollo%20Agropecuario%2C%20Rural%2C%20Forestal%20y%20Pesquero&param_formato=20&param_unidad=1&param_tipo=10&param_lenguaje=1&param_clasificacion=32&"
    ],
    "sheet":[
        "1.4.1S",
        "1.4.2S",
        "1.4.3S",
        "1.4.4S",
        "1.4.5S",
        "1.4.6S",
        "1.4.7S",
        "1.4.8S",
        "1.4.9S",
        "1.4.10S",
        #"1.4.11S",
        #Lo dejaron de actualizar en 2016
        "1.4.12S",
        "1.4.13S",
        "1.4.14S",
        "1.4.15S",
        "1.4.16S",
        "1.4.17S",
        "1.4.18S",
        "1.4.19S",
        "1.4.20S",
        "1.4.21S",
        "1.4.22S",
        "1.4.23S",
        "1.4.24S",
        "1.4.25S",
        "1.4.26S",
        "1.4.27S",
        "1.4.28S",
        "1.4.29S",
        "1.4.30S",
        "1.4.31S",
        "1.4.32S",
        "1.4.33S",
        #"1.4.34S",
        #Lo dejaron de actualizar en 2016
        #"1.4.35S",
        #Lo dejaron de actualizar en 2016
        #"1.4.36S",
        #Lo dejaron de actualizar en 2016
        #"1.4.37S",
        #Lo dejaron de actualizar en 2016
        #"1.4.38S",
        #Lo dejaron de actualizar en 2016
        "1.4.39S",
        "1.4.40S",
        "1.4.41S",
        "1.4.42S",
        "1.4.43S",
        "1.4.44S",
        "1.4.45S",
        "1.4.46S",
        "1.4.47S",
        "1.4.48S",
        "1.4.49S",
        "1.4.50S",
        "1.4.51S",
        "1.4.52S",
        "1.4.53S",
        "1.4.54S",
        "1.4.55S",
        "1.4.56S",
        "1.4.57S",
        "1.4.58S"
    ],
    "titulos":[
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        #False,
        #Lo dejaron de actualizar en 2016
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        #False,
        #Lo dejaron de actualizar en 2016
        #False,
        #Lo dejaron de actualizar en 2016
        #False,
        #Lo dejaron de actualizar en 2016
        #False,
        #Lo dejaron de actualizar en 2016
        #False,
        #Lo dejaron de actualizar en 2016
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False
    ]
}

for key in list(app_dict.keys()):
    print(f'{key}: {len(app_dict[key])}')

executable_path = {'executable_path': 'chromedriver.exe'}
browser = Browser('chrome', **executable_path, headless=False)
base_url = "http://presto.hacienda.gob.mx/presto/files/system/mashlets/app_layout_estopor/index.html?param_formato_desc="

df_dict = {
    "title":[],
    "df":[],
    "sheet":[]
}

for i in range(0,len(app_dict["title"])):
    df = scrape(i,browser,app_dict)
    title = browser.find_by_css(".cls-title-table").find_by_css(".cls-title-main").text
    df_dict["title"].append(title)
    df_dict["df"].append(df)
    df_dict["sheet"].append(app_dict["sheet"][i])



to_download = store(df_dict)

scraped_to_official()

if to_download:
    print("\n\n\n")
    print("IMPORTANT: please download the following tables by hand...")
    for table in to_download:
        print(table)
    print(f"SHCP inserted a new column in these tables. Download by hand and replace in test.xlsm with exact same format as other tables (column numbers on top, headers on row 3), save, and execute this script again.")
              

