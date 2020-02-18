import requests
import pymongo
import pandas as pd
import time
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import load_workbook
from splinter import Browser
import win32com.client

def click_on_something(strings,types_,by_num='',by_text = '',used_already = []):
    found = False
    my_dict = {}
    my_dict["strings"] = strings
    my_dict["types"] = types_
    #print(my_dict)
    j = 0
    while found == False:
        try: 
            to_click = browser
            for i in range(0,len(my_dict["strings"])):
                if my_dict["types"][i] == "id":
                    to_click = to_click.find_by_id(my_dict["strings"][i])
                elif my_dict["types"][i] == "css":
                    blah = f".{my_dict['strings'][i]}"
                    to_click = to_click.find_by_css(blah)
                    #In the case of selecting boundlists, must select the one that hasent been used yet
                    if used_already and my_dict['strings'][i] == 'x-boundlist':
                        for obj in to_click:
                            if obj not in used_already:
                                to_click = obj
                else:
                    print("types_ invalid input: must be id or css")
                #print(f"to click: {to_click}")
            #If not looking to click on the first returned object
            if not by_num:
                if by_text:
                    clicked = False
                    for obj in to_click:
                        if obj.text == by_text:
                            obj.click()
                            clicked = True
                            break
                    if clicked == False:
                        print("by_text invalid input: text not found")
                        
                else:
                    to_click.click()
                found = True
            #if looking to click on object by number
            elif by_num:
                to_click[int(by_num)].click()
                found = True
            else:
                print("by_num invalid input: must be empty or an integer")
                break
            return to_click
        
        except AttributeError:
            j += 1
            print(f"looking for element. attempt number {j}")
            
        if j > 10:
            print("Either app did't load properly, or internet is down/weak")
            break
            

def scrape(num_,browser,app_dict):
    browser.visit(f'{base_url}{app_dict["url"][num_]}')
    
    #Clicking the serie drop_down
    loaded = False
    while loaded == False:
        try:
            click_on_something(["tab-1013"],["id"])
            try:
                series_drop = click_on_something(["cmp_series_SERIE-triggerWrap","x-trigger-index-0"],["id","css"])
                loaded = True
            except:
                click_on_something(["tab-1012"],["id"])
                print("walao")
        except:
            pass

    #Clicking the todas las series button
    #select_all = browser.find_by_css(".list-ct").find_by_css(".x-column-header-text").first
    #select_all.click()
    series_clicked = False
    while series_clicked == False:
        click_on_something(["list-ct","x-column-header-text"],["css","css"],0)
        test = browser.find_by_id("cmp_series_SERIE-inputEl")._element.get_attribute('value')
        if test:
            series_clicked = True
                  
    #Unclicking series drop_down
    series_drop.click()

    #Clicking on cifras drop down
    #cifras_drop = browser.find_by_id("cmp_cuadro_PRESENTACION").find_by_css(".x-trigger-index-0")
    #cifras_drop.click()
    cifras_drop_clicked = False
    while cifras_drop_clicked == False:
        try:
            cifras_drop = click_on_something(["cmp_cuadro_PRESENTACION","x-trigger-index-0"],["id","css"])
            cifras_drop_clicked = True
        except:
            print("cifras drop not found yet...")



    #Clicking on millones de pesos
    cifras_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],False,"Millones de pesos")

    #Upon clicking the dropdown, an x-boundlist is generated, must add it to used_aleady to we dont use it again
    used_already = []
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year lower limit drop down
    #low_lim_drop = browser.find_by_id("cmp_cuadro_DE").find_by_css(".x-trigger-cell").find_by_css(".x-trigger-index-0")
    #low_lim_drop.click()
    low_lim_clicked = False
    while low_lim_clicked == False:
        try:
            low_lim_drop = click_on_something(["cmp_cuadro_DE","x-trigger-cell","x-trigger-index-0"],["id","css","css"])
            low_lim_clicked = True
        except:
            print("year lower limit drop down not found yet...")

    #Clicking on the second to latest year (hence the 1 in the click_on_something function)in the low_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],1,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Clicking on year upper limit drop down
    #up_lim_drop = browser.find_by_id("cmp_cuadro_A").find_by_css(".x-trigger-index-0")
    #up_lim_drop.click()
    up_lim_clicked = False
    while up_lim_clicked == False:
        try:
            up_lim_drop = click_on_something(["cmp_cuadro_A","x-trigger-index-0"],["id","css"])
            up_lim_clicked = True
        except:
            print("up_lim_drop not found yet...")
            
    #Clicking on the latest year in the up_lim_opts option menu
    low_lim_opts = click_on_something(["x-boundlist","x-boundlist-item"],["css","css"],0,'',used_already)
    #adding bound list generated by clicking low_lim_drop to used_already
    used_already = add_to_used(used_already,browser.find_by_css(".x-boundlist"))
    used_already

    #Consultar Series (clicking the button and suppressing warnings until the page loads the data)
    #consultar_series = browser.find_by_id("cmp_series_CONSULTAR_SERIES-btnEl")
    #consultar_series.click()
    success = False
    while success == False:
        clicked = False
        try:
            consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
            clicked = True
            print("consultar button clicked")
        except:
            print("no consultar button")
        if clicked == True:
            try:
                alert = browser.get_alert()
                alert.accept()
                print("alert accepted")
            except:
                print("there is no alert")
                time.sleep(0.5)
                test = browser.find_by_css(".x-mask-msg")
                #Sometimes the button isnt clicked, thus...
                print(f'test = {test}')
                if test == []:
                    consultar_series = click_on_something(["cmp_series_CONSULTAR_SERIES-btnEl"],["id"])
                    print("had to click again")
                success = True

    loaded = False

    while loaded == False:
        try:
            the_table = browser.find_by_id("GridSeries")
            header_row = the_table.find_by_css(".x-box-inner")[1]
            loaded = True
        except:
            print("loading...")
            #sometimes the consultar button isnt clicked in the previous step, thus...
                
            
    #Finding Headers of the table
    header_divs = header_row.find_by_css(".x-unselectable")
    idx = 0
    headers = []
    for header in header_divs:
        if idx == 0:
            headers.append(header.find_by_tag("span").find_by_tag("label")._element.get_attribute("innerHTML").strip())
            idx += 1
        else:
            headers.append(header.find_by_tag("span")._element.get_attribute("innerHTML").replace("\n","").replace("\r","").strip())

    #Scraping the data and creating a data frame
    table_large = the_table.find_by_id("GridSeries-body")

    table = table_large.find_by_tag("table")

    rows = table.find_by_tag("tr")

    data_rows = list(rows)[1:]

    the_dict = dict((el,[]) for el in headers)

    for row in data_rows:
        to_append = row.text.split("\n")
        for k in range(0,len(the_dict)):
            the_dict[headers[k]].append(to_append[k]) 

    df =pd.DataFrame(the_dict)
    return df

def store(df_dict):
    #Loading the excel file
    print("Loading test.xlsm...")
    wb = load_workbook(filename='N:/Lsolis/eco/SHCP/Historicos/test.xlsm', read_only=False, keep_vba=True)
    idx = 0
    for worksheet in df_dict["sheet"]:
        #Loading the sheet
        sheet_ranges = wb[worksheet]
        
        #determining whether or not a new row(s) must be inserted
        done = False
        row_ = 3
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                #print(sheet_ranges[f'A{row_}'].value)
                row_ += 1
            else:
                done = True

        last_row = row_ - 1

        current_year = int(sheet_ranges[f'A{last_row}'].value[-4:])
                  
        #current month is the month of the lates data already archived
        #Here, current_trim is defined incase this table is in quarters instead of months
        current_month = int(sheet_ranges[f'A{last_row}'].value[:2])
        current_trim = current_month
        
        #this takes care of case where data is being scraped during last month of the year (12)
        #or last quarter of the year (04)
        #current_year_scraped is the year of the latest data that was just scraped.
        current_year_scraped = int(df_dict["df"][idx].iloc[-1,0].split("/")[1])
                  
        if "Mensual" in df_dict["df"][idx].columns[0]:
            if current_month == 12 and current_year_scraped > current_year:
                current_year += 1
        elif "Trimestral" in df_dict["df"][idx].columns[0]:
            if current_trim == 4 and current_year_scraped > current_year:
                current_year += 1
        else:
            print("error with time headers")
            break
                  
        if current_year == current_year_scraped:
            previous_2year = current_year - 2
        elif (current_year_scraped - current_year) == 1:
            previous_2year = current_year - 1
        else:
            print("Please download this table by hand from SHCP, replace in test2, and try again.")
            break

        previous_year = current_year - 1

        #This finds the nuber of rows of data so far in the 2 years
        done = False
        row_ = 4
        while done == False:
            if sheet_ranges[f'A{row_}'].value or sheet_ranges[f'A{row_}'].value == 0:
                if int(sheet_ranges[f'A{row_}'].value[-4:]) == previous_2year:
                    last_row_2previous = row_
                    #print(sheet_ranges[f'A{row_}'].value)  
                row_ += 1
            else:
                done = True
                  
        #insert rows if there are rows to be inserted
        if len(df_dict["df"][idx]) > (last_row-last_row_2previous):
            sheet_ranges.insert_rows(last_row + 1, len(df_dict["df"][idx])-(last_row - last_row_2previous))
            print(f'{len(df_dict["df"][idx])-(last_row-last_row_2previous)} rows were inserted.')
        
        #finding row range to which values will be inserted
        ins_begin = last_row_2previous + 1
        ins_end = ins_begin + len(df_dict["df"][idx]) - 1
        
        
        #Making sure headers in excel are equal to headers in dataframe
        sheet_col_heads = []
        sheet_col_heads2 = []
        for col in range(0,count_columns(sheet_ranges)):
            sheet_col_heads.append(sheet_ranges['A3'].offset(0,col).value.replace("_x000D_",""))
            sheet_col_heads2.append(sheet_ranges['A3'].offset(0,col).value)
              
        stripped_col_heads = []
        for head in sheet_col_heads:
            stripped_col_heads.append(head.strip())
            
        stripped_col_heads2 = []
        for head in sheet_col_heads2:
            stripped_col_heads2.append(head.strip())

        compare = list(df_dict["df"][idx].columns)
        compare2 = []
        
        for header in compare:
            compare2.append(header.replace("\r","").replace("\n","").strip())
        
        print(f'{df_dict["sheet"][idx]}: {stripped_col_heads == compare2}')
        
        #comparing the sheet columns to the scraped columns
        if stripped_col_heads == compare2:
            pass
        else:
            print(f"sheet column headers don't agree with scraped scraped column headers")
            break
            
            
        #Insert Values from dataframe into excel with some format
        row_height = sheet_ranges.row_dimensions[last_row_2previous].height
        for col in range(1,len(stripped_col_heads) + 1):
            for rw in range(ins_begin, ins_end + 1):
                #pasting value in cell
                try:
                    sheet_ranges.cell(rw,col).value = float((df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",",""))
                    sheet_ranges.cell(rw,col).number_format = '#,###.0'
                except ValueError:
                    sheet_ranges.cell(rw,col).value = (df_dict["df"][idx][stripped_col_heads[col - 1]][rw-ins_begin]).replace(",","")
                #setting font style and size    
                sheet_ranges.cell(rw,col).font = Font(name='Arial', size=9)
                #setting border of cell
                border = Border(left=Side(border_style='thin', color='E4E4E4'),
                                right=Side(border_style='thin', color='E4E4E4'),
                                top=Side(border_style='thin', color='E4E4E4'),
                                bottom=Side(border_style='thin', color='E4E4E4'))
                sheet_ranges.cell(rw,col).border = border
                #setting alignment of text in cell
                alignment=Alignment(horizontal='right',
                                    vertical='top')
                sheet_ranges.cell(rw,col).alignment = alignment
                #setting row height
                sheet_ranges.row_dimensions[rw].height = row_height
                #setting fill color of cell
                greyFill = PatternFill(start_color='DCDCDC',
                           end_color='DCDCDC',
                           fill_type='solid')
                whiteFill = PatternFill(start_color='FFFFFF',
                    end_color='FFFFFF',
                    fill_type='solid')
                if rw % 2 == 0:
                    sheet_ranges.cell(rw,col).fill = greyFill 
                else:
                    sheet_ranges.cell(rw,col).fill = whiteFill
            print(f"column {col} out of {len(stripped_col_heads)} inserted successfully")
        
        idx += 1
    print("Saving test.xlsm...")
    wb.save(filename = 'N:/Lsolis/eco/SHCP/Historicos/test.xlsm')
    print("Data was stored in test.xlsm")

def add_to_used(used_already,xboundlists):
    for lst in xboundlists:
        el_id = lst._element.get_attribute('id')
        if el_id not in used_already:
            used_already.append(el_id)
    return used_already

def count_columns(sheet_ranges):
    non_empty = True
    col_ = 1
    col_count = 0
    while non_empty:
        if sheet_ranges['A3'].offset(0,col_ - 1).value:
            col_ +=1
            col_count += 1
        else:
            non_empty = False
    return col_count

def scraped_to_official():
              
    print("Copying data from test to official file...")
    xl=win32com.client.Dispatch("Excel.Application")
    
    print(" Loading official file...")
    wb = xl.Workbooks.Open("N:/Lsolis/eco/SHCP/Historicos/1Ingreso_gasto_fincanc_sector_publico.xlsm")
    
    print(" Running macro...")
    xl.Application.Run("1Ingreso_gasto_fincanc_sector_publico.xlsm!CopySheetsFromScrapedData")

    wb.Save()
    wb.Close()

    del xl
    print("Finished!")
                  
app_dict = {
    "title":[
        "Balance del Sector Público",
        "Situación Financiera del Sector Público",
        "Ahorro Corriente del Sector Público Presupuestario",
        "Financiamiento de la Inversión Física y Financiera Presupuestaria",
        "Financiamiento del Balance Público por Institución",
        "Financiamiento del Balance Público por Sector",
        "Ingresos Presupuestarios del Sector Público",
        "Ingresos Petroleros del Sector Público",
        "Ingresos Presupuestarios del Gobierno Federal, Petroleros y No petroleros",
        "Ingresos Presupuestarios del Gobierno Federal (Resumen Artículo 1 LIF)",
        "Ingresos Propios de Organismos y Empresas Bajo Control Presupuestario Directo",
        "Gasto Neto del Sector Público Presupuestario",
        "Gasto programable del sector público presupuestario",
        "Gasto Programable del Sector Público Presupuestario. Clasificación Funcional",
        "Gasto Programable del Sector Público Presupuestario. Clasificación Funcional-Económica",
        "Servicios Personales del Sector Público Presupuestario. Clasificación Funcional",
        "Otros Gastos de Operación del Sector Público Presupuestario. Clasificación Funcional",
        "Subsidios, Transferencias y Aportaciones del Sector Público Presupuestario distintos de Servicios Personales. Clasificación Funcional",
        "Ayudas y Otros Gastos Corrientes del Sector Público Presupuestario. Clasificación Funcional",
        "Inversión Física del Sector Público Presupuestario. Clasificación Funcional",
        "Subsidios, Transferencias y Aportaciones del Sector Público Presupuestario Incluyendo Servicios Personales. Clasificación Funcional",
        "Inversión física presupuestaria",
        "Costo Financiero del Sector Público Presupuestario",
        "Costo Financiero del Sector Público Presupuestario. Clasificación Institucional",
        "Gasto del Gobierno Federal en Clasificación Económica",
        "Principales Beneficiarios de Transferencias del Gobierno Federal. Clasificación Funcional",
        "Participaciones a Entidades Federativas y Municipios por Fondo",
        "Adeudos de Ejercicios Fiscales Anteriores (ADEFAS). Clasificación Económica",
        "Adeudos de Ejercicios Fiscales Anteriores (ADEFAS). Clasificación Administrativa"
    ],
    "url":[
        "Balance%20del%20Sector%20P%C3%BAblico&param_formato=1&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Situaci%C3%B3n%20Financiera%20del%20Sector%20P%C3%BAblico&param_formato=2&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Ahorro%20Corriente%20del%20Sector%20P%C3%BAblico%20Presupuestario&param_formato=18&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Financiamiento%20de%20la%20Inversi%C3%B3n%20F%C3%ADsica%20y%20Financiera%20Presupuestaria&param_formato=19&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Financiamiento%20del%20Balance%20P%C3%BAblico%20por%20Instituci%C3%B3n&param_formato=20&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Financiamiento%20del%20Balance%20P%C3%BAblico%20por%20Sector&param_formato=21&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Ingresos%20Presupuestarios%20del%20Sector%20P%C3%BAblico&param_formato=3&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Ingresos%20Petroleros%20del%20Sector%20P%C3%BAblico&param_formato=29&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Ingresos%20Presupuestarios%20del%20Gobierno%20Federal%2C%20Petroleros%20y%20No%20petroleros&param_formato=2&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Ingresos%20Presupuestarios%20del%20Gobierno%20Federal%20%20(Resumen%20Art%C3%ADculo%201%20LIF)&param_formato=3&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Ingresos%20Propios%20de%20Organismos%20y%20Empresas%20Bajo%20Control%20Presupuestario%20Directo&param_formato=4&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=3&",
        "Gasto%20Neto%20del%20Sector%20P%C3%BAblico%20Presupuestario&param_formato=5&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Gasto%20programable%20del%20sector%20p%C3%BAblico%20presupuestario&param_formato=15&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Gasto%20Programable%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional&param_formato=8&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Gasto%20Programable%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional-Econ%C3%B3mica&param_formato=9&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Servicios%20Personales%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional&param_formato=23&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Otros%20Gastos%20de%20Operaci%C3%B3n%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional&param_formato=24&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Subsidios%2C%20Transferencias%20y%20Aportaciones%20del%20Sector%20P%C3%BAblico%20Presupuestario%20distintos%20de%20Servicios%20Personales.%20Clasificaci%C3%B3n%20Funcional&param_formato=25&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Ayudas%20y%20Otros%20Gastos%20Corrientes%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional&param_formato=26&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Inversi%C3%B3n%20F%C3%ADsica%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Funcional&param_formato=27&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Subsidios%2C%20Transferencias%20y%20Aportaciones%20del%20Sector%20P%C3%BAblico%20Presupuestario%20Incluyendo%20Servicios%20Personales.%20Clasificaci%C3%B3n%20Funcional&param_formato=28&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Inversi%C3%B3n%20f%C3%ADsica%20presupuestaria&param_formato=11&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Costo%20Financiero%20del%20Sector%20P%C3%BAblico%20Presupuestario&param_formato=12&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Costo%20Financiero%20del%20Sector%20P%C3%BAblico%20Presupuestario.%20Clasificaci%C3%B3n%20Institucional&param_formato=13&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=1&",
        "Gasto%20del%20Gobierno%20Federal%20en%20Clasificaci%C3%B3n%20Econ%C3%B3mica&param_formato=29&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Principales%20Beneficiarios%20de%20Transferencias%20del%20Gobierno%20Federal.%20Clasificaci%C3%B3n%20Funcional&param_formato=10&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Participaciones%20a%20Entidades%20Federativas%20y%20Municipios%20por%20Fondo&param_formato=17&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Adeudos%20de%20Ejercicios%20Fiscales%20Anteriores%20(ADEFAS).%20Clasificaci%C3%B3n%20Econ%C3%B3mica&param_formato=14&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&",
        "Adeudos%20de%20Ejercicios%20Fiscales%20Anteriores%20(ADEFAS).%20Clasificaci%C3%B3n%20Administrativa&param_formato=15&param_unidad=1&param_tipo=1&param_lenguaje=1&param_clasificacion=2&"
    ],
    "sheet":[
        "1.1.1S",
        "1.1.2S",
        "1.1.3S",
        "1.1.4S",
        "1.1.5S",
        "1.1.6S",
        "1.2.1S",
        "1.2.2S",
        "1.2.3S",
        "1.2.4S",
        "1.2.5S",
        "1.3.1S",
        "1.3.2S",
        "1.3.3S",
        "1.3.4S",
        "1.3.5S",
        "1.3.6S",
        "1.3.7S",
        "1.3.8S",
        "1.3.9S",
        "1.3.10S",
        "1.3.11S",
        "1.3.12S",
        "1.3.13S",
        "1.3.14S",
        "1.3.15S",
        "1.3.16S",
        "1.3.17S",
        "1.3.18S"
    ]
}

#Opening chrome with chromedriver
executable_path = {'executable_path': 'chromedriver.exe'}
browser = Browser('chrome', **executable_path, headless=False)

base_url = "http://presto.hacienda.gob.mx/presto/files/system/mashlets/app_layout_estopor/index.html?param_formato_desc="


df_dict = {
    "title":[],
    "df":[],
    "sheet":[]
}
                  
for i in range(0,len(app_dict["title"])):
    df = scrape(i,browser,app_dict)
    title = browser.find_by_css(".cls-title-table").find_by_css(".cls-title-main").text
    df_dict["title"].append(title)
    df_dict["df"].append(df)
    df_dict["sheet"].append(app_dict["sheet"][i])

print("All the data has been scraped.")

store(df_dict)

scraped_to_official()
              